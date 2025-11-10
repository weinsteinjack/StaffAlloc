"""Excel export helpers for StaffAlloc reports."""

from __future__ import annotations

import io
from datetime import date
from typing import Dict, Iterable, List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app import models


def _auto_size(ws) -> None:
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
        adjusted = min(length + 4, 60)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = adjusted


def _header(row) -> None:
    header_font = Font(bold=True, color="FFFFFF")
    fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    for cell in row:
        cell.font = header_font
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center")


def portfolio_workbook(*, projects: Iterable[models.Project], over_allocated: List, bench: List) -> io.BytesIO:
    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Portfolio"

    ws_summary.append([
        "Project",
        "Manager",
        "Status",
        "Funded Hours",
        "Allocated Hours",
        "Utilization %",
        "Start",
        "Sprints",
    ])
    _header(ws_summary[1])

    for project in projects:
        funded = sum(a.funded_hours for a in project.assignments)
        allocated = sum(
            allocation.allocated_hours
            for assignment in project.assignments
            for allocation in assignment.allocations or []
        )
        utilization = (allocated / funded * 100) if funded else 0
        ws_summary.append([
            project.name,
            project.manager.full_name if project.manager else "—",
            project.status,
            funded,
            allocated,
            round(utilization, 2),
            project.start_date.isoformat(),
            project.sprints,
        ])

    _auto_size(ws_summary)

    ws_over = wb.create_sheet("Over-allocated")
    ws_over.append(["Employee", "Role", "FTE %", "Projects"])
    _header(ws_over[1])
    for employee in over_allocated:
        projects_str = ", ".join(
            f"{p.project_name} ({p.allocated_hours}h)" for p in getattr(employee, "projects", [])
        )
        ws_over.append([
            employee.full_name,
            getattr(employee, "role", None) or "—",
            round(employee.fte_percentage, 2),
            projects_str,
        ])
    _auto_size(ws_over)

    ws_bench = wb.create_sheet("Bench")
    ws_bench.append(["Employee", "Role", "FTE %", "Available Hours"])
    _header(ws_bench[1])
    for employee in bench:
        ws_bench.append([
            employee.full_name,
            getattr(employee, "role", None) or "—",
            round(employee.fte_percentage, 2),
            getattr(employee, "available_hours", 0),
        ])
    _auto_size(ws_bench)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def project_workbook(project: models.Project) -> io.BytesIO:
    wb = Workbook()
    ws_info = wb.active
    ws_info.title = "Project"
    ws_info.append(["Name", "Code", "Client", "Status", "Start", "Sprints", "Manager"])
    _header(ws_info[1])
    ws_info.append([
        project.name,
        project.code,
        project.client or "—",
        project.status,
        project.start_date.isoformat(),
        project.sprints,
        project.manager.full_name if project.manager else "—",
    ])
    _auto_size(ws_info)

    ws_assignments = wb.create_sheet("Assignments")
    ws_assignments.append([
        "Employee",
        "Role",
        "LCAT",
        "Funded Hours",
        "Allocated Hours",
        "Remaining Hours",
    ])
    _header(ws_assignments[1])

    assignment_map = {}
    for assignment in project.assignments:
        allocated = sum(a.allocated_hours for a in assignment.allocations or [])
        remaining = assignment.funded_hours - allocated
        ws_assignments.append([
            assignment.user.full_name,
            assignment.role.name,
            assignment.lcat.name,
            assignment.funded_hours,
            allocated,
            remaining,
        ])
        assignment_map[assignment.id] = assignment
    _auto_size(ws_assignments)

    ws_allocations = wb.create_sheet("Allocations")
    ws_allocations.append([
        "Employee",
        "Year",
        "Month",
        "Allocated Hours",
    ])
    _header(ws_allocations[1])

    for assignment in project.assignments:
        for allocation in assignment.allocations or []:
            ws_allocations.append([
                assignment.user.full_name,
                allocation.year,
                allocation.month,
                allocation.allocated_hours,
            ])
    _auto_size(ws_allocations)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

