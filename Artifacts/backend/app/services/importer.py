"""Excel import helpers for project data."""

from __future__ import annotations

import datetime as dt
import io
import logging
from typing import Dict, List, Optional, Tuple

from openpyxl import load_workbook

from app import crud, models, schemas
from app.services.ai import (
    GeminiConfigurationError,
    GeminiInvocationError,
    suggest_header_mapping,
)


class ProjectImportError(Exception):
    """Raised when an import file cannot be processed."""


logger = logging.getLogger(__name__)


def _normalize_header(value: Optional[str]) -> Optional[str]:
    return value.strip().lower() if isinstance(value, str) else None


def _row_dict(row, header_map) -> Dict[str, Optional[str]]:
    result: Dict[str, Optional[str]] = {}
    for key, idx in header_map.items():
        cell = row[idx] if idx < len(row) else None
        result[key] = cell.value if cell is not None else None
    return result


def parse_projects_sheet(sheet) -> List[Dict[str, str]]:
    required_headers = {
        'name': {'name', 'project', 'project name'},
        'code': {'code', 'project code'},
        'client': {'client', 'customer'},
        'start_date': {'start date', 'start'},
        'sprints': {'sprints', 'duration (sprints)'},
        'status': {'status'}
    }
    header_cells = next(sheet.iter_rows(min_row=1, max_row=1))
    header_map = _resolve_header_map("Projects", header_cells, required_headers)

    missing = [field for field in required_headers if field not in header_map]
    if missing:
        raise ProjectImportError(f"Projects sheet is missing columns: {', '.join(missing)}")

    projects: List[Dict[str, str]] = []
    for row in sheet.iter_rows(min_row=2):
        values = _row_dict(row, header_map)
        if not values['name'] or not values['code']:
            continue
        projects.append({
            'name': str(values['name']).strip(),
            'code': str(values['code']).strip(),
            'client': str(values['client']).strip() if values['client'] else None,
            'start_date': str(values['start_date']).strip(),
            'sprints': str(values['sprints']).strip(),
            'status': str(values['status']).strip() if values['status'] else 'Active'
        })
    return projects


def parse_assignments_sheet(sheet) -> List[Dict[str, str]]:
    headers = {
        'project_code': {'project code', 'code'},
        'employee_email': {'employee email', 'email', 'user email'},
        'role': {'role', 'job role'},
        'lcat': {'lcat', 'labor category'},
        'funded_hours': {'funded hours', 'funded'}
    }
    header_cells = next(sheet.iter_rows(min_row=1, max_row=1))
    header_map = _resolve_header_map("Assignments", header_cells, headers)

    missing = [field for field in headers if field not in header_map]
    if missing:
        raise ProjectImportError(f"Assignments sheet is missing columns: {', '.join(missing)}")

    assignments: List[Dict[str, str]] = []
    for row in sheet.iter_rows(min_row=2):
        values = _row_dict(row, header_map)
        if not values['project_code'] or not values['employee_email']:
            continue
        assignments.append({
            'project_code': str(values['project_code']).strip(),
            'employee_email': str(values['employee_email']).strip().lower(),
            'role': str(values['role']).strip() if values['role'] else 'Unassigned',
            'lcat': str(values['lcat']).strip() if values['lcat'] else 'General',
            'funded_hours': str(values['funded_hours']).strip() if values['funded_hours'] else '0'
        })
    return assignments


def parse_allocations_sheet(sheet) -> List[Dict[str, str]]:
    headers = {
        'project_code': {'project code', 'code'},
        'employee_email': {'employee email', 'email', 'user email'},
        'year': {'year'},
        'month': {'month'},
        'hours': {'hours', 'allocated hours'}
    }
    header_cells = next(sheet.iter_rows(min_row=1, max_row=1))
    header_map = _resolve_header_map("Allocations", header_cells, headers)

    missing = [field for field in headers if field not in header_map]
    if missing:
        raise ProjectImportError(f"Allocations sheet is missing columns: {', '.join(missing)}")

    allocations: List[Dict[str, str]] = []
    for row in sheet.iter_rows(min_row=2):
        values = _row_dict(row, header_map)
        if not values['project_code'] or not values['employee_email']:
            continue
        allocations.append({
            'project_code': str(values['project_code']).strip(),
            'employee_email': str(values['employee_email']).strip().lower(),
            'year': str(values['year']).strip(),
            'month': str(values['month']).strip(),
            'hours': str(values['hours']).strip()
        })
    return allocations


def _parse_date(value: str) -> dt.date:
    for fmt in ('%Y-%m-%d', '%m/%d/%Y', '%d/%m/%Y'):
        try:
            return dt.datetime.strptime(value, fmt).date()
        except ValueError:
            continue
    # fallback to Excel serial number
    try:
        serial = float(value)
        base_date = dt.date(1899, 12, 30)
        return base_date + dt.timedelta(days=int(serial))
    except Exception as exc:
        raise ProjectImportError(f"Unable to parse date value '{value}'") from exc


def _parse_int(value: str, field: str) -> int:
    try:
        return int(float(value))
    except (TypeError, ValueError) as exc:
        raise ProjectImportError(f"Unable to parse integer for {field}: '{value}'") from exc


def _resolve_header_map(sheet_name: str, header_cells, required_headers: Dict[str, set[str]]) -> Dict[str, int]:
    header_lookup: Dict[str, int] = {}
    header_map: Dict[str, int] = {}

    for idx, cell in enumerate(header_cells):
        header_value = _normalize_header(cell.value)
        if not header_value:
            continue
        header_lookup[header_value] = idx
        for field, aliases in required_headers.items():
            if field in header_map:
                continue
            if header_value in aliases:
                header_map[field] = idx

    missing = [field for field in required_headers if field not in header_map]
    if not missing:
        return header_map

    headers_as_list = [str(cell.value or "").strip() for cell in header_cells]
    try:
        ai_mapping = suggest_header_mapping(
            headers=headers_as_list,
            required_fields=missing,
            sheet_name=sheet_name,
        )
    except GeminiConfigurationError:
        logger.debug(
            "Gemini header mapping unavailable; continuing with deterministic mapping",
            exc_info=True,
        )
        return header_map
    except GeminiInvocationError as exc:
        logger.warning("Gemini header mapping failed for %s sheet: %s", sheet_name, exc)
        return header_map

    for field, header_name in ai_mapping.items():
        if field in header_map:
            continue
        normalized = _normalize_header(header_name)
        if not normalized:
            continue
        idx = header_lookup.get(normalized)
        if idx is not None:
            header_map[field] = idx

    return header_map


def import_projects_from_workbook(
    *,
    data: bytes,
    db,
    manager_id: Optional[int]
) -> Tuple[List[models.Project], List[str]]:
    try:
        workbook = load_workbook(io.BytesIO(data))
    except Exception as exc:
        raise ProjectImportError("Uploaded file is not a valid Excel workbook") from exc

    sheet_map = {name.lower(): name for name in workbook.sheetnames}

    if 'projects' not in sheet_map:
        raise ProjectImportError("Workbook must contain a 'Projects' sheet")

    assignments_sheet = workbook[sheet_map.get('assignments')] if 'assignments' in sheet_map else None
    allocations_sheet = workbook[sheet_map.get('allocations')] if 'allocations' in sheet_map else None

    projects_data = parse_projects_sheet(workbook[sheet_map['projects']])
    assignments_data = parse_assignments_sheet(assignments_sheet) if assignments_sheet else []
    allocations_data = parse_allocations_sheet(allocations_sheet) if allocations_sheet else []

    created_projects: List[models.Project] = []
    skipped_codes: List[str] = []
    project_lookup: Dict[str, models.Project] = {}

    for record in projects_data:
        start_date = _parse_date(record['start_date'])
        sprints = _parse_int(record['sprints'], 'sprints')
        payload = schemas.ProjectCreate(
            name=record['name'],
            code=record['code'],
            client=record['client'],
            start_date=start_date,
            sprints=sprints,
            status=record['status'],
            manager_id=manager_id
        )
        existing = crud.get_project_by_code(db, code=payload.code, manager_id=manager_id)
        if existing:
            project_lookup[payload.code] = existing
            skipped_codes.append(payload.code)
            continue
        project = crud.create_project(db, project=payload)
        project_lookup[payload.code] = project
        created_projects.append(project)

    db.flush()

    # Refresh to include relationships
    for code, project in project_lookup.items():
        project_lookup[code] = crud.get_project(db, project_id=project.id)

    # Process assignments
    assignment_lookup: Dict[Tuple[int, int], models.ProjectAssignment] = {}
    for record in assignments_data:
        project = project_lookup.get(record['project_code'])
        if not project:
            raise ProjectImportError(f"Unknown project code '{record['project_code']}' in Assignments sheet")

        user = crud.get_user_by_email(db, email=record['employee_email'])
        if not user:
            raise ProjectImportError(f"Unknown employee email '{record['employee_email']}'")

        role = crud.get_role_by_name(db, name=record['role'], owner_id=manager_id)
        if not role:
            role = crud.create_role(db, schemas.RoleCreate(name=record['role']), owner_id=manager_id)

        lcat = crud.get_lcat_by_name(db, name=record['lcat'], owner_id=manager_id)
        if not lcat:
            lcat = crud.create_lcat(db, schemas.LCATCreate(name=record['lcat']), owner_id=manager_id)

        funded_hours = _parse_int(record['funded_hours'], 'funded_hours')
        assignment = crud.get_assignment_by_user_and_project(db, user_id=user.id, project_id=project.id)
        if not assignment:
            assignment = crud.create_project_assignment(
                db,
                schemas.ProjectAssignmentCreate(
                    project_id=project.id,
                    user_id=user.id,
                    role_id=role.id,
                    lcat_id=lcat.id,
                    funded_hours=funded_hours,
                ),
            )
        else:
            crud.update_project_assignment(
                db,
                assignment_id=assignment.id,
                assignment_update=schemas.ProjectAssignmentUpdate(
                    role_id=role.id,
                    lcat_id=lcat.id,
                    funded_hours=funded_hours,
                ),
            )
        assignment_lookup[(project.id, user.id)] = assignment

    db.flush()

    # Process allocations
    for record in allocations_data:
        project = project_lookup.get(record['project_code'])
        if not project:
            raise ProjectImportError(f"Unknown project code '{record['project_code']}' in Allocations sheet")

        user = crud.get_user_by_email(db, email=record['employee_email'])
        if not user:
            raise ProjectImportError(f"Unknown employee email '{record['employee_email']}' in Allocations sheet")

        assignment = assignment_lookup.get((project.id, user.id))
        if not assignment:
            assignment = crud.get_assignment_by_user_and_project(db, user_id=user.id, project_id=project.id)
            if not assignment:
                raise ProjectImportError(
                    f"No assignment for {record['employee_email']} on project {record['project_code']}"
                )

        year = _parse_int(record['year'], 'year')
        month = _parse_int(record['month'], 'month')
        hours = _parse_int(record['hours'], 'hours')

        existing = next(
            (a for a in assignment.allocations or [] if a.year == year and a.month == month),
            None,
        )
        if existing:
            crud.update_allocation(
                db,
                allocation_id=existing.id,
                allocation_update=schemas.AllocationUpdate(allocated_hours=hours),
            )
        else:
            crud.create_allocation(
                db,
                schemas.AllocationCreate(
                    project_assignment_id=assignment.id,
                    year=year,
                    month=month,
                    allocated_hours=hours,
                ),
            )

    db.commit()

    # Reload projects with fresh relationships
    refreshed = [crud.get_project(db, project_id=project.id) for project in created_projects]
    return [project for project in refreshed if project], skipped_codes

